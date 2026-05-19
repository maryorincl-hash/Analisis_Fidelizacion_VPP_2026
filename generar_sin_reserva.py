import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

HOY = pd.Timestamp('2026-05-19')

# ── 1. Cargar leads ──────────────────────────────────────────────────────────

archivos_leads = {
    'Enero':   r'C:\Users\maryorin.vivas\Downloads\Enero_leads.xlsx',
    'Febrero': r'C:\Users\maryorin.vivas\Downloads\Febrero_leads.xlsx',
    'Marzo':   r'C:\Users\maryorin.vivas\Downloads\Marzo_leads.xlsx',
    'Abril':   r'C:\Users\maryorin.vivas\Downloads\Abril_leads.xlsx',
    'Mayo':    r'C:\Users\maryorin.vivas\Downloads\Mayo_leads.xlsx',
}
COL_LEADS = {
    0:'fecha_lead', 1:'responsable', 2:'cliente', 3:'email',
    4:'marca', 5:'modelo', 6:'origen', 7:'sub_origen',
    8:'empresa', 9:'tipo_negocio', 10:'etapa_lead', 11:'creado_por',
    12:'gestionado_ia', 13:'no_contactable_wp', 14:'rut',
    15:'telefono', 16:'patente'
}

dfs = []
for mes, path in archivos_leads.items():
    d = pd.read_excel(path, header=0)
    d.columns = range(len(d.columns))
    d = d.rename(columns=COL_LEADS)
    d['mes'] = mes
    dfs.append(d)

leads = pd.concat(dfs, ignore_index=True)
leads['patente']    = leads['patente'].astype(str).str.strip().str.upper().replace('NAN', pd.NA)
leads['fecha_lead'] = pd.to_datetime(leads['fecha_lead'], errors='coerce')
leads['origen']     = leads['origen'].astype(str).str.strip()

def clasificar_canal(origen):
    o = str(origen).upper()
    if 'PROPIO'     in o: return 'Sitio Web Propio'
    elif 'MARKETPLACE' in o: return 'Marketplace'
    elif 'SALON'    in o: return 'Salón'
    elif 'FACEBOOK' in o or 'RRSS' in o: return 'RRSS / Facebook'
    elif 'ADS'      in o or 'GOOGLE' in o: return 'Campañas Ads'
    elif 'CONTACT'  in o: return 'Contact Center'
    elif 'WHATSAPP' in o or 'WA' in o: return 'WhatsApp'
    else: return 'Otro'

leads['canal'] = leads['origen'].apply(clasificar_canal)

# ── 2. Cargar oportunidades ──────────────────────────────────────────────────

opp_raw = pd.read_excel(r'C:\Users\maryorin.vivas\Downloads\opportunity_19_05_2026_11_24.xlsx', header=0)
opp_raw.columns = range(len(opp_raw.columns))
COL_OPP = {
    0:'id', 1:'etapa', 2:'patente', 3:'empresa', 4:'marca',
    5:'fecha_reserva', 6:'fecha_facturacion', 7:'cod_sucursal',
    8:'precio', 9:'creado_por', 10:'origen', 11:'id_inspeccion',
    12:'nombre', 13:'pct_descuento', 14:'margen', 15:'vpp',
    16:'financiamiento', 17:'fecha_creacion_opp', 18:'sucursal',
    19:'rut', 20:'telefono', 21:'email'
}
opp = opp_raw.rename(columns=COL_OPP)
opp['patente']       = opp['patente'].astype(str).str.strip().str.upper().replace('NAN', pd.NA)
opp['fecha_reserva'] = pd.to_datetime(opp['fecha_reserva'], errors='coerce')

# ── 3. Tabla de reservas ─────────────────────────────────────────────────────

reservas = (
    opp[opp['fecha_reserva'].notna() & opp['patente'].notna()]
    .groupby('patente')
    .agg(fecha_reserva=('fecha_reserva','min'))
    .reset_index()
)

# ── 4. Cruzar leads con patente vs reservas ──────────────────────────────────

leads_pat = leads[leads['patente'].notna()].copy()
leads_cruzado = leads_pat.merge(reservas, on='patente', how='left')
leads_cruzado['estado'] = leads_cruzado['fecha_reserva'].apply(
    lambda fr: 'Sin reserva' if pd.isna(fr) else 'Con reserva'
)

sin_reserva = leads_cruzado[leads_cruzado['estado'] == 'Sin reserva'].copy()

# ── 5. Agregar por patente ───────────────────────────────────────────────────

fecha_30d = HOY - pd.Timedelta(days=30)

leads_30d = (
    sin_reserva[sin_reserva['fecha_lead'] >= fecha_30d]
    .groupby('patente').size().reset_index(name='leads_30d')
)

canal_principal = (
    sin_reserva.groupby(['patente','canal']).size()
    .reset_index(name='n')
    .sort_values('n', ascending=False)
    .groupby('patente').first()
    .reset_index()[['patente','canal']]
    .rename(columns={'canal':'canal_principal'})
)

agrupado = (
    sin_reserva.groupby('patente').agg(
        marca        = ('marca',      lambda x: x.mode().iloc[0] if len(x.mode()) else ''),
        modelo       = ('modelo',     lambda x: x.mode().iloc[0] if len(x.mode()) else ''),
        primer_lead  = ('fecha_lead', 'min'),
        ultimo_lead  = ('fecha_lead', 'max'),
        total_leads  = ('patente',    'count'),
    ).reset_index()
)

agrupado = agrupado.merge(canal_principal, on='patente', how='left')
agrupado = agrupado.merge(leads_30d, on='patente', how='left')
agrupado['leads_30d'] = agrupado['leads_30d'].fillna(0).astype(int)
agrupado['dias_vitrina'] = (HOY - agrupado['primer_lead']).dt.days.fillna(0).astype(int)

# Ordenar: patentes más antiguas primero
agrupado = agrupado.sort_values('primer_lead', ascending=True).reset_index(drop=True)

# ── 6. Clasificación y acción sugerida ───────────────────────────────────────

def clasificar(row):
    dias  = row['dias_vitrina']
    leads = row['total_leads']
    if dias < 30:
        return 'Reciente'
    elif leads > 50 and dias > 60:
        return 'Alta tracción sin reserva'
    elif leads > 50:
        return 'Alta tracción'
    elif leads >= 10:
        return 'Tracción moderada'
    else:
        return 'Baja tracción'

def accion(row):
    c = row['clasificacion']
    if   c == 'Alta tracción sin reserva': return 'Revisar precio o condición — alta demanda sin conversión'
    elif c == 'Alta tracción':             return 'Monitorear — puede necesitar ajuste de precio'
    elif c == 'Tracción moderada':         return 'Validar competitividad de precio en vitrina'
    elif c == 'Baja tracción':             return 'Revisar publicación y fotografías'
    else:                                  return 'En período de exposición inicial'

agrupado['clasificacion']   = agrupado.apply(clasificar, axis=1)
agrupado['accion_sugerida'] = agrupado.apply(accion,     axis=1)

print(f'[OK] Patentes sin reserva: {len(agrupado):,}')
print(f'     Alta traccion sin reserva: {(agrupado["clasificacion"]=="Alta traccion sin reserva").sum()}')
print(f'     Baja traccion: {(agrupado["clasificacion"]=="Baja traccion").sum()}')

# ── 7. Generar Excel ─────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = 'Sin Reserva - Por Patente'

# Estilos
AZUL_OSC   = 'FF1F3A6E'
AZUL_MED   = 'FF2E75B6'
NARANJA    = 'FFED7D31'
ROJO       = 'FFC00000'
VERDE      = 'FF70AD47'
AMARILLO   = 'FFFFC000'
GRIS_FOND  = 'FFF5F5F5'
BLANCO     = 'FFFFFFFF'

h_font  = Font(name='Calibri', bold=True, color=BLANCO, size=11)
h_fill  = PatternFill('solid', fgColor=AZUL_OSC)
h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin    = Side(style='thin', color='FFCCCCCC')
borde   = Border(left=thin, right=thin, top=thin, bottom=thin)

# Color por clasificación
COLOR_CLAS = {
    'Alta tracción sin reserva': ROJO,
    'Alta tracción':             NARANJA,
    'Tracción moderada':         AMARILLO,
    'Baja tracción':             'FFD9D9D9',
    'Reciente':                  VERDE,
}

# ── Fila de título ──
ws.merge_cells('A1:K1')
titulo = ws['A1']
titulo.value     = 'Patentes sin Reserva — Unidades Publicadas sin Conversión | Salazar-Israel Enero–Mayo 2026'
titulo.font      = Font(name='Calibri', bold=True, color=BLANCO, size=13)
titulo.fill      = PatternFill('solid', fgColor=AZUL_OSC)
titulo.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# ── Fila de resumen KPIs ──
ws.merge_cells('A2:C2'); ws['A2'].value = f'Total patentes sin reserva: {len(agrupado):,}'
ws.merge_cells('D2:F2'); ws['D2'].value = f'Alta tracción sin reserva: {(agrupado["clasificacion"]=="Alta tracción sin reserva").sum()}'
ws.merge_cells('G2:I2'); ws['G2'].value = f'Baja tracción (+30 días): {(agrupado["clasificacion"]=="Baja tracción").sum()}'
ws.merge_cells('J2:K2'); ws['J2'].value = f'Generado: {HOY.strftime("%d/%m/%Y")}'
for col in ['A','D','G','J']:
    ws[f'{col}2'].font      = Font(name='Calibri', bold=True, color=AZUL_MED, size=10)
    ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col}2'].fill      = PatternFill('solid', fgColor='FFE8F0FB')
ws.row_dimensions[2].height = 20

# ── Encabezados ──
HEADERS = [
    'Patente', 'Marca', 'Modelo',
    'Primer Lead', 'Último Lead', 'Días en Vitrina',
    'Total Leads', 'Leads Últimos 30d', 'Canal Principal',
    'Clasificación', 'Acción Sugerida'
]
WIDTHS = [12, 14, 28, 14, 14, 14, 12, 16, 20, 24, 52]

for col_idx, (header, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font      = h_font
    cell.fill      = h_fill
    cell.alignment = h_align
    cell.border    = borde
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[3].height = 32

# ── Filas de datos ──
for i, (_, row) in enumerate(agrupado.iterrows(), start=4):
    clas       = row['clasificacion']
    color_clas = COLOR_CLAS.get(clas, 'FFD9D9D9')
    fondo      = PatternFill('solid', fgColor=GRIS_FOND if i % 2 == 0 else BLANCO)

    valores = [
        row['patente'],
        row['marca'],
        row['modelo'],
        row['primer_lead'].strftime('%d/%m/%Y') if pd.notna(row['primer_lead']) else '',
        row['ultimo_lead'].strftime('%d/%m/%Y') if pd.notna(row['ultimo_lead']) else '',
        row['dias_vitrina'],
        row['total_leads'],
        row['leads_30d'],
        row['canal_principal'] if pd.notna(row['canal_principal']) else '',
        clas,
        row['accion_sugerida'],
    ]

    for col_idx, valor in enumerate(valores, start=1):
        cell = ws.cell(row=i, column=col_idx, value=valor)
        cell.border    = borde
        cell.alignment = Alignment(horizontal='center' if col_idx not in [3, 11] else 'left',
                                   vertical='center', wrap_text=(col_idx == 11))
        cell.font      = Font(name='Calibri', size=10)
        if col_idx == 10:
            cell.fill = PatternFill('solid', fgColor=color_clas)
            cell.font = Font(name='Calibri', size=10, bold=True,
                             color=BLANCO if clas in ('Alta tracción sin reserva', 'Alta tracción') else '00000000')
        elif col_idx == 1:
            cell.font = Font(name='Calibri', size=10, bold=True, color=AZUL_MED)
            cell.fill = fondo
        else:
            cell.fill = fondo

    ws.row_dimensions[i].height = 18

# ── Congelar encabezado ──
ws.freeze_panes = 'A4'

# ── Filtros automáticos ──
ws.auto_filter.ref = f'A3:K{3 + len(agrupado)}'

# ── Segunda hoja: Leyenda ────────────────────────────────────────────────────
ws2 = wb.create_sheet('Leyenda')
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 58
ws2.column_dimensions['C'].width = 30

leyenda_titulo = ws2['A1']
leyenda_titulo.value     = 'Leyenda — Clasificación de Patentes'
leyenda_titulo.font      = Font(name='Calibri', bold=True, color=BLANCO, size=12)
leyenda_titulo.fill      = PatternFill('solid', fgColor=AZUL_OSC)
leyenda_titulo.alignment = Alignment(horizontal='center', vertical='center')
ws2.merge_cells('A1:C1')
ws2.row_dimensions[1].height = 26

leyenda_headers = ['Clasificación', 'Criterio', 'Acción Sugerida']
for col_idx, h in enumerate(leyenda_headers, 1):
    c = ws2.cell(row=2, column=col_idx, value=h)
    c.font = Font(name='Calibri', bold=True, color=BLANCO, size=10)
    c.fill = PatternFill('solid', fgColor=AZUL_MED)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = borde

leyenda_data = [
    ('Alta tracción sin reserva', 'Más de 50 leads y más de 60 días publicada sin reserva',
     'Revisar precio o condición — alta demanda sin conversión', ROJO, BLANCO),
    ('Alta tracción',             'Más de 50 leads y menos de 60 días publicada',
     'Monitorear — puede necesitar ajuste de precio', NARANJA, BLANCO),
    ('Tracción moderada',         'Entre 10 y 50 leads recibidos',
     'Validar competitividad de precio en vitrina', AMARILLO, '00000000'),
    ('Baja tracción',             'Menos de 10 leads y más de 30 días publicada',
     'Revisar publicación y fotografías', 'FFD9D9D9', '00000000'),
    ('Reciente',                  'Menos de 30 días desde el primer lead',
     'En período de exposición inicial', VERDE, BLANCO),
]
for row_idx, (clas, criterio, accion_txt, color, txt_color) in enumerate(leyenda_data, 3):
    for col_idx, valor in enumerate([clas, criterio, accion_txt], 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=valor)
        c.border    = borde
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.font      = Font(name='Calibri', size=10,
                           bold=(col_idx==1), color=txt_color if col_idx==1 else '00000000')
        c.fill      = PatternFill('solid', fgColor=color) if col_idx == 1 else PatternFill('solid', fgColor=BLANCO)
    ws2.row_dimensions[row_idx].height = 30

# ── Guardar ──────────────────────────────────────────────────────────────────
SALIDA = r'C:\Users\maryorin.vivas\Downloads\sin_reserva_vitrina.xlsx'
wb.save(SALIDA)
print(f'[OK] Excel generado: {SALIDA}')
print(f'     Filas de datos: {len(agrupado):,}')
